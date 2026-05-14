#!/usr/bin/env python3
"""
Build VOGDB ↔ ICTV taxonomy mapping for virus phylogeny evaluation.

For each VOGDB family, extracts virus species/host names from FAA headers,
matches them to ICTV Master Species List taxonomy using genus-level fuzzy matching,
and saves a mapping pickle for downstream monophyly evaluation.

Output:
  --output-map:  pickle dict {vfam_id: {seq_name: (genus, family, order, class)}}
  --output-stats: text file with coverage statistics
"""
import os, re, sys, csv, pickle, argparse
from collections import defaultdict

def build_ictv_lookup(msl_path):
    """Parse ICTV MSL xlsx into genus → (genus, family, order, class) lookup."""
    import openpyxl
    wb = openpyxl.load_workbook(msl_path)
    ws = wb['MSL']

    genus_taxonomy = {}  # genus_lower -> (genus, family, order, class)
    for r in range(2, ws.max_row + 1):
        genus = ws.cell(r, 14).value
        family = ws.cell(r, 12).value
        order = ws.cell(r, 10).value
        class_ = ws.cell(r, 8).value
        if genus and genus.strip():
            key = genus.strip().lower()
            if key not in genus_taxonomy:
                genus_taxonomy[key] = (genus.strip(), family or '', order or '', class_ or '')

    # Sort by length descending (greedy longest-first matching)
    genus_sorted = sorted(genus_taxonomy.keys(), key=len, reverse=True)
    return genus_taxonomy, genus_sorted


def match_vogdb_to_ictv(name, genus_taxonomy, genus_sorted):
    """
    Fast multi-strategy matcher. Uses string split + containment, no regex.
    Returns (match_type, taxonomy_tuple or None).
    """
    name = name.strip().lower()
    if not name:
        return 'no_match', None

    # 1. Exact match
    if name in genus_taxonomy:
        return 'exact', genus_taxonomy[name]

    # 2. Check whole words (split on spaces, hyphens, underscores)
    name_words = set(name.replace('-', ' ').replace('_', ' ').split())
    for genus_lower in genus_sorted:
        if genus_lower in name_words:
            return 'word', genus_taxonomy[genus_lower]

    # 3. Check contiguous substrings for longer genus names (>=6 chars)
    for genus_lower in genus_sorted:
        if len(genus_lower) >= 6 and genus_lower in name:
            return 'substring', genus_taxonomy[genus_lower]

    return 'no_match', None


# Global cache for species name matching
_MATCH_CACHE = {}

def match_with_cache(name, genus_taxonomy, genus_sorted):
    if name in _MATCH_CACHE:
        return _MATCH_CACHE[name]
    result = match_vogdb_to_ictv(name, genus_taxonomy, genus_sorted)
    _MATCH_CACHE[name] = result
    return result


def parse_species_labels(faa_path):
    """Extract species/host strings from FAA file headers."""
    labels = {}
    with open(faa_path) as f:
        for line in f:
            if line.startswith('>'):
                name = line[1:].split()[0]
                matches = re.findall(r'\[([^\]]+)\]', line)
                labels[name] = matches[-1].strip() if matches else None
    return labels


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--msl-path', default='virus_data/ictv_msl.xlsx',
                        help='Path to ICTV MSL xlsx')
    parser.add_argument('--faa-dir', default='virus_data/faa',
                        help='Directory with VOGDB FAA files')
    parser.add_argument('--output-map', default='virus_data/vogdb_ictv_map.pickle',
                        help='Output pickle: {vfam_id: {seq: (genus,family,order,class)}}')
    parser.add_argument('--output-stats', default='eval_preds/vogdb_ictv_stats.txt',
                        help='Output coverage statistics')
    parser.add_argument('--max-families', type=int, default=0,
                        help='Limit for testing (0=all)')
    parser.add_argument('--family-list', default='',
                        help='Optional: only process families from this ref pickle (faster)')
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    msl_path = os.path.join(script_dir, args.msl_path)
    faa_dir = os.path.join(script_dir, args.faa_dir)
    os.makedirs(os.path.dirname(os.path.join(script_dir, args.output_map)), exist_ok=True)
    os.makedirs(os.path.dirname(os.path.join(script_dir, args.output_stats)), exist_ok=True)

    print("=" * 60)
    print("VOGDB ↔ ICTV Taxonomy Mapper")
    print("=" * 60)
    print(f"ICTV MSL: {msl_path}")
    print(f"FAA dir:  {faa_dir}")

    # 1. Build ICTV lookup
    print("\n[1/4] Building ICTV taxonomy lookup...")
    genus_taxonomy, genus_sorted = build_ictv_lookup(msl_path)
    print(f"  ICTV genera:  {len(genus_taxonomy)}")
    families = set(v[1] for v in genus_taxonomy.values() if v[1])
    print(f"  ICTV families: {len(families)}")

    # 2. Scan FAA files (possibly filtered by family list)
    import glob
    print("\n[2/4] Scanning VOGDB FAA files...")

    # If --family-list is provided, only process those families
    if args.family_list:
        family_list_path = os.path.join(script_dir, args.family_list)
        ref_data = pickle.load(open(family_list_path, 'rb'))
        target_vfam_ids = set(ref_data.keys())
        print(f"  Family list from: {args.family_list} ({len(target_vfam_ids)} families)")
    else:
        target_vfam_ids = None

    faa_files = sorted(glob.glob(os.path.join(faa_dir, 'VFAM*.faa')))
    if args.max_families > 0:
        faa_files = faa_files[:args.max_families]

    # Filter to only target families if specified
    if target_vfam_ids:
        faa_files = [f for f in faa_files
                     if os.path.basename(f).replace('.faa', '') in target_vfam_ids]
    print(f"  FAA files to process: {len(faa_files)}")

    vfam_to_taxonomy = {}  # {vfam_id: {seq_name: (genus, family, order, class)}}
    stats = defaultdict(int)
    match_type_stats = defaultdict(int)
    n_matched_seqs = 0
    n_total_seqs = 0

    for idx, faa_path in enumerate(faa_files):
        vfam_id = os.path.basename(faa_path).replace('.faa', '')
        labels = parse_species_labels(faa_path)
        seq_tax = {}
        for seq_name, host_str in labels.items():
            n_total_seqs += 1
            if host_str is None:
                seq_tax[seq_name] = None
                continue
            mt, tax = match_with_cache(host_str, genus_taxonomy, genus_sorted)
            match_type_stats[mt] += 1
            if tax:
                seq_tax[seq_name] = tax
                n_matched_seqs += 1
            else:
                seq_tax[seq_name] = None

        vfam_to_taxonomy[vfam_id] = seq_tax

        # Count genera per family
        genera_in_fam = set()
        families_in_fam = set()
        for tax in seq_tax.values():
            if tax:
                genera_in_fam.add(tax[0])
                families_in_fam.add(tax[1])
        if len(genera_in_fam) >= 2:
            stats['multi_genus'] += 1
        if len(families_in_fam) >= 2:
            stats['multi_family'] += 1
        if len(genera_in_fam) >= 1:
            stats['any_match'] += 1
        stats['total_families'] += 1

        if (idx + 1) % 4000 == 0:
            print(f"  Progress: {idx+1}/{len(faa_files)}")

    # 3. Print statistics
    print("\n[3/4] Matching statistics:")
    total = sum(match_type_stats.values())
    for mt in ['exact', 'word_boundary', 'substring', 'no_match']:
        cnt = match_type_stats.get(mt, 0)
        print(f"  {mt:>15}: {cnt:>7} ({cnt/total*100:.1f}%)" if total else f"  {mt:>15}: {cnt}")

    print(f"\n  Total sequences:     {n_total_seqs}")
    print(f"  Matched sequences:   {n_matched_seqs} ({n_matched_seqs/n_total_seqs*100:.1f}%)" if n_total_seqs else "")
    nf = stats['total_families']
    print(f"\n  Total families:      {nf}")
    print(f"  Any match:           {stats.get('any_match',0)} ({stats.get('any_match',0)/nf*100:.1f}%)")
    print(f"  Multi-genus:         {stats.get('multi_genus',0)} ({stats.get('multi_genus',0)/nf*100:.1f}%)")
    print(f"  Multi-family:        {stats.get('multi_family',0)} ({stats.get('multi_family',0)/nf*100:.1f}%)")

    # 4. Save
    print(f"\n[4/4] Saving output...")
    with open(os.path.join(script_dir, args.output_map), 'wb') as f:
        pickle.dump(vfam_to_taxonomy, f)
    print(f"  Map pickle: {args.output_map} ({len(vfam_to_taxonomy)} families)")

    stats_path = os.path.join(script_dir, args.output_stats)
    with open(stats_path, 'w') as f:
        f.write("VOGDB ↔ ICTV Taxonomy Mapping Statistics\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"Total VOGDB families:           {nf}\n")
        f.write(f"Families with any match:        {stats.get('any_match',0)} ({stats.get('any_match',0)/nf*100:.1f}%)\n")
        f.write(f"Families multi-genus (>=2):     {stats.get('multi_genus',0)} ({stats.get('multi_genus',0)/nf*100:.1f}%)\n")
        f.write(f"Families multi-family (>=2):    {stats.get('multi_family',0)} ({stats.get('multi_family',0)/nf*100:.1f}%)\n\n")
        f.write(f"Total sequences:  {n_total_seqs}\n")
        f.write(f"Matched sequences: {n_matched_seqs} ({n_matched_seqs/n_total_seqs*100:.1f}%)\n\n")
        f.write("Match type breakdown:\n")
        total = sum(match_type_stats.values()) or 1
        for mt in ['exact', 'word_boundary', 'substring', 'no_match']:
            cnt = match_type_stats.get(mt, 0)
            f.write(f"  {mt:>15}: {cnt:>7} ({cnt/total*100:.1f}%)\n")

    print(f"  Stats text: {stats_path}")
    print("\nDone!")


if __name__ == '__main__':
    main()
