#!/usr/bin/env python3
"""
Build ICTV taxonomy mapping and assess VOGDB coverage for monophyly evaluation.
Runs on CPU, ~5-10 minutes for 5000 families.
"""
import openpyxl, re, os, glob, pickle, sys
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

print("="*60)
print("ICTV-VOGDB Taxonomy Mapping Assessment")
print("="*60)

# 1. Build ICTV taxonomy lookup
print("\n[1/4] Building ICTV taxonomy lookup...")
wb = openpyxl.load_workbook(os.path.join(SCRIPT_DIR, 'virus_data/ictv_msl.xlsx'))
ws = wb['MSL']

genus_taxonomy = {}  # genus_lower -> (genus, family, order, class)

for r in range(2, ws.max_row + 1):
    genus = ws.cell(r, 14).value
    family = ws.cell(r, 12).value
    order = ws.cell(r, 10).value
    class_ = ws.cell(r, 8).value
    if genus and genus.strip():
        key = genus.strip().lower()
        if key not in genus_taxonomy:
            genus_taxonomy[key] = (genus.strip(), family or '', order or '', class_ or '')

genus_sorted = sorted(genus_taxonomy.keys(), key=len, reverse=True)
print(f"  ICTV genera: {len(genus_taxonomy)}")
print(f"  ICTV families: {len(set(v[1] for v in genus_taxonomy.values() if v[1]))}")

# 2. Define matcher
print("\n[2/4] Defining taxonomy matcher...")

def match_vogdb_to_ictv(vogdb_name):
    """Match a VOGDB species name to ICTV taxonomy genus."""
    name = vogdb_name.strip().lower()
    if name in genus_taxonomy:
        return genus_taxonomy[name]
    # Word-boundary match: genus as whole word
    for gl in genus_sorted:
        if re.search(r'\b' + re.escape(gl) + r'\b', name):
            return genus_taxonomy[gl]
    # Substring match for genera >= 6 chars (with greedy longest-first)
    for gl in genus_sorted:
        if len(gl) >= 6 and gl in name:
            return genus_taxonomy[gl]
    return None

# 3. Process all VOGDB families
print("\n[3/4] Processing VOGDB families...")
faa_dir = os.path.join(SCRIPT_DIR, 'virus_data/faa')
vogdb_species_by_family = defaultdict(set)

faa_files = sorted(glob.glob(os.path.join(faa_dir, 'VFAM*.faa')))
total_fams = len(faa_files)
print(f"  Total VOGDB families: {total_fams}")

for i, f in enumerate(faa_files):
    vid = os.path.basename(f).replace('.faa', '')
    with open(f) as fh:
        for line in fh:
            if line.startswith('>'):
                m = re.findall(r'\[([^\]]+)\]', line)
                if m:
                    vogdb_species_by_family[vid].add(m[-1].strip())
    if (i + 1) % 5000 == 0:
        print(f"    Processed {i+1}/{total_fams} families")

print(f"  Done. Unique species strings: {sum(len(v) for v in vogdb_species_by_family.values())}")

# 4. Compute coverage statistics
print("\n[4/4] Computing coverage statistics...")
stats = {
    'total_families': len(vogdb_species_by_family),
    'families_1plus_genus': 0,     # >=1 sequence matched to genus
    'families_2plus_genera': 0,    # >=2 distinct genera (monophyly testable)
    'families_2plus_families': 0,  # >=2 distinct families
    'families_2plus_orders': 0,    # >=2 distinct orders
}

family_taxonomy = {}  # vid -> {sequence_name -> (genus, family, order, class)}

for vid, species_set in vogdb_species_by_family.items():
    seq_tax = {}
    genera = set()
    families = set()
    orders = set()
    
    for sp in species_set:
        tax = match_vogdb_to_ictv(sp)
        if tax:
            seq_tax[sp] = tax
            genera.add(tax[0])
            families.add(tax[1])
            orders.add(tax[2])
    
    if len(genera) >= 1:
        stats['families_1plus_genus'] += 1
    if len(genera) >= 2:
        stats['families_2plus_genera'] += 1
    if len(families) >= 2:
        stats['families_2plus_families'] += 1
    if len(orders) >= 2:
        stats['families_2plus_orders'] += 1
    
    if len(genera) >= 2 and len(seq_tax) >= 4:
        family_taxonomy[vid] = seq_tax

# Report
n = stats['total_families']
print(f"\n{'='*60}")
print(f"  ICTV Monophyly Evaluation Feasibility")
print(f"{'='*60}")
print(f"  Total VOGDB families:           {n}")
print(f"  >=1 seq matched to ICTV:        {stats['families_1plus_genus']} ({stats['families_1plus_genus']/n*100:.1f}%)")
print(f"  >=2 ICTV genera (monophyly):    {stats['families_2plus_genera']} ({stats['families_2plus_genera']/n*100:.1f}%)")
print(f"  >=2 ICTV families:              {stats['families_2plus_families']} ({stats['families_2plus_families']/n*100:.1f}%)")
print(f"  >=2 ICTV orders:                {stats['families_2plus_orders']} ({stats['families_2plus_orders']/n*100:.1f}%)")
print(f"")
print(f"  Families saved for evaluation:  {len(family_taxonomy)}")
print(f"  (>=2 genera AND >=4 matched sequences)")

# Save taxonomy mapping
out_path = os.path.join(SCRIPT_DIR, 'virus_data', 'vogdb_ictv_taxonomy.pickle')
pickle.dump(family_taxonomy, open(out_path, 'wb'))
print(f"\n  Taxonomy mapping saved: {out_path}")
print(f"  File size: {os.path.getsize(out_path)/1024/1024:.1f} MB")
print(f"{'='*60}")
